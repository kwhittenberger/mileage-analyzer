#!/usr/bin/env python3
"""
Mileage Analysis Tool for D'Ewart Representatives, L.L.C.
Analyzes trip data and categorizes business vs personal mileage
"""

import csv
import sys
from datetime import datetime, timedelta
from collections import defaultdict
import re
import time
import json
import os
import webbrowser
import urllib.parse

try:
    from geopy.geocoders import Nominatim
    from geopy.exc import GeocoderTimedOut, GeocoderServiceError
    GEOPY_AVAILABLE = True
except ImportError:
    GEOPY_AVAILABLE = False

try:
    import googlemaps
    GOOGLE_MAPS_AVAILABLE = True
except ImportError:
    GOOGLE_MAPS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Configuration - Key addresses (loaded from config.json)
HOME_ADDRESS = "15815 61st Ln NE, Kenmore"  # Default
WORK_ADDRESS = "9227 NE 180th St, Bothell"  # Default
COUPEVILLE_AREA = ["Coupeville", "Oak Harbor", "Clinton", "Whidbey"]

# Distance threshold for business trips (miles)
BUSINESS_DISTANCE_THRESHOLD = 8.0

# Known business location keywords
BUSINESS_KEYWORDS = [
    "gas", "station", "fuel", "shell", "chevron", "76", "arco", "bp",
    "costco", "safeway", "store", "market", "shop",
    "customer", "client", "office"
]

def normalize_address(address):
    """Normalize address for comparison"""
    if not address:
        return ""
    return address.strip().lower()

def is_home_address(address):
    """Check if address is home"""
    norm = normalize_address(address)
    norm_home = normalize_address(HOME_ADDRESS)
    # Fuzzy match - check if key parts of home address are in the given address
    return norm_home in norm or norm in norm_home

def is_work_address(address):
    """Check if address is work"""
    norm = normalize_address(address)
    norm_work = normalize_address(WORK_ADDRESS)
    # Fuzzy match - check if key parts of work address are in the given address
    return norm_work in norm or norm in norm_work

def is_bothell_area(address):
    """Check if address is in Bothell area"""
    norm = normalize_address(address)
    return "bothell" in norm or "kenmore" in norm

def is_whidbey_area(address):
    """Check if address is on Whidbey Island"""
    norm = normalize_address(address)
    for location in COUPEVILLE_AREA:
        if location.lower() in norm:
            return True
    return False

def is_portland_area(address):
    """Check if address is in Portland area"""
    norm = normalize_address(address)
    portland_keywords = ["portland", " or ", "oregon", "beaverton", "tigard", "gresham", "hillsboro"]
    return any(keyword in norm for keyword in portland_keywords)

def is_spokane_area(address):
    """Check if address is in Spokane area"""
    norm = normalize_address(address)
    spokane_keywords = ["spokane", "spokane valley", "liberty lake"]
    return any(keyword in norm for keyword in spokane_keywords)

def extract_street_info(address):
    """Extract street number and street name from address"""
    norm = normalize_address(address)
    parts = norm.split()

    street_number = None
    street_name_parts = []

    # Try to find street number (first numeric part)
    for i, part in enumerate(parts):
        if part.replace('-', '').isdigit():
            street_number = int(part.replace('-', ''))
            # Rest is street name
            street_name_parts = parts[i+1:]
            break

    # Join street name (exclude city/state/zip)
    # Remove common endings
    street_name = ' '.join(street_name_parts)
    for ending in [' wa', ' washington', ' or', ' oregon', ' seattle', ' bothell', ' kenmore', ' woodinville']:
        if street_name.endswith(ending):
            street_name = street_name[:-len(ending)]
            break

    return street_number, street_name.strip()

def find_nearby_addresses(address, resolved_addresses, max_number_diff=100):
    """Find resolved addresses that are nearby (same street, close numbers)"""
    street_num, street_name = extract_street_info(address)

    if not street_num or not street_name:
        return []

    nearby = []

    for resolved_addr, business_name in resolved_addresses.items():
        resolved_num, resolved_street = extract_street_info(resolved_addr)

        if not resolved_num or not resolved_street:
            continue

        # Check if same street (fuzzy match)
        if resolved_street in street_name or street_name in resolved_street:
            # Check if numbers are close
            num_diff = abs(street_num - resolved_num)
            if num_diff <= max_number_diff and num_diff > 0:  # Don't match exact same number
                nearby.append({
                    'address': resolved_addr,
                    'business_name': business_name,
                    'distance': num_diff
                })

    # Sort by distance (closest first)
    nearby.sort(key=lambda x: x['distance'])

    return nearby

def is_business_location(address):
    """Check if address appears to be a business location"""
    norm = normalize_address(address)
    for keyword in BUSINESS_KEYWORDS:
        if keyword in norm:
            return True
    return False

# Business mapping file - unified storage for all address->business mappings
# Format: {address: {name: str, category: str, source: str}}
# source can be "manual" or "google_api"
BUSINESS_MAPPING_FILE = "business_mapping.json"
CONFIG_FILE = "config.json"
business_mapping = {}
google_api_key = None
gmaps_client = None


def load_business_mapping():
    """Load business name mappings from file

    Handles multiple formats for backward compatibility:
    - Old format: {address: name}
    - New format: {address: {name, category, source}}
    """
    global business_mapping
    if os.path.exists(BUSINESS_MAPPING_FILE):
        try:
            with open(BUSINESS_MAPPING_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Filter out comment fields
                business_mapping = {k: v for k, v in data.items() if not k.startswith('_')}
        except:
            business_mapping = {}


def save_business_mapping():
    """Save business mappings to file"""
    try:
        with open(BUSINESS_MAPPING_FILE, 'w', encoding='utf-8') as f:
            json.dump(business_mapping, f, indent=2, ensure_ascii=False)
    except:
        pass


def get_mapping_name(address: str) -> str:
    """Get the business name from mapping, handling both old and new format"""
    value = business_mapping.get(address)
    if value is None:
        return None
    if isinstance(value, dict):
        return value.get('name', '')
    return value


def get_mapping_category(address: str) -> str:
    """Get the category from mapping if available (new format only)"""
    value = business_mapping.get(address)
    if isinstance(value, dict):
        return value.get('category')
    return None


def get_mapping_source(address: str) -> str:
    """Get the source of a mapping (manual or google_api)"""
    value = business_mapping.get(address)
    if isinstance(value, dict):
        return value.get('source', 'manual')
    return 'manual'


def set_mapping_entry(address: str, name: str, category: str = None, source: str = "manual"):
    """Set a business mapping entry"""
    entry = {'name': name, 'source': source}
    if category:
        entry['category'] = category
    business_mapping[address] = entry


def load_config():
    """Load configuration including Google API key and addresses"""
    global google_api_key, gmaps_client, HOME_ADDRESS, WORK_ADDRESS
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                api_key = config.get('google_places_api_key', '').strip()
                if api_key and GOOGLE_MAPS_AVAILABLE:
                    google_api_key = api_key
                    gmaps_client = googlemaps.Client(key=api_key)

                # Load home and work addresses if configured
                home = config.get('home_address', '').strip()
                if home:
                    HOME_ADDRESS = home

                work = config.get('work_address', '').strip()
                if work:
                    WORK_ADDRESS = work
        except:
            pass

def lookup_business_google_places(address):
    """Look up business using Google Places API"""
    if not gmaps_client or not address:
        return None

    try:
        # Method 1: Try geocoding the address first, then search nearby
        # This is more accurate for finding businesses at specific addresses
        try:
            geocode_result = gmaps_client.geocode(address)
            if geocode_result and len(geocode_result) > 0:
                location = geocode_result[0]['geometry']['location']
                lat, lng = location['lat'], location['lng']

                # Search for places within 25 meters of this exact address
                nearby_result = gmaps_client.places_nearby(
                    location=(lat, lng),
                    radius=25,  # 25 meters - very close to the address
                    rank_by=None
                )

                if nearby_result and nearby_result.get('status') == 'OK' and nearby_result.get('results'):
                    for place in nearby_result['results']:
                        name = place.get('name', '')
                        types = place.get('types', [])

                        # Filter out non-business types
                        unwanted = ['street_address', 'premise', 'subpremise', 'route', 'political', 'locality']
                        business_types = [t for t in types if t not in unwanted]

                        if name and business_types:
                            print(f"    ✓ FOUND BUSINESS (nearby): {name} at {address[:40]}")
                            sys.stdout.flush()
                            return name
        except Exception as nearby_error:
            # If nearby search fails, fall back to text search
            pass

        # Method 2: Fall back to text search if nearby search didn't work
        result = gmaps_client.places(query=address)

        if result and result.get('status') == 'OK' and result.get('results'):
            # Get the first (most relevant) result
            place = result['results'][0]

            # Extract business name
            name = place.get('name', '')

            # Check if this is actually a business or just an address coming back
            # If the name is essentially the same as the query address, it's not a real business
            normalized_name = name.lower().replace(',', '').replace('.', '')
            normalized_address = address.lower().replace(',', '').replace('.', '')

            # Check if name is just the address reformatted
            if normalized_name in normalized_address or normalized_address in normalized_name:
                print(f"    No business found (just address) at {address[:40]}")
                sys.stdout.flush()
                return None

            # Get business type for better context
            types = place.get('types', [])

            # Filter out generic/unwanted types that indicate it's not a real business
            unwanted = ['street_address', 'premise', 'subpremise', 'route', 'political', 'locality']
            business_types = [t for t in types if t not in unwanted]

            # If there are no actual business types, it's probably just an address
            if not business_types:
                print(f"    No business types found at {address[:40]}")
                sys.stdout.flush()
                return None

            # If we have a name and it's not just a street address
            if name and not any(word in name.lower() for word in ['street', 'avenue', 'road', 'lane', 'drive', ' st ', ' ave ', ' rd ', ' ln ', ' dr ']):
                print(f"    ✓ FOUND BUSINESS: {name} at {address[:40]}")
                sys.stdout.flush()
                return name

            # If name seems generic, add business type context
            if name and business_types:
                first_type = business_types[0].replace('_', ' ').title()
                if first_type.lower() not in name.lower():
                    result_name = f"{name} ({first_type})"
                    print(f"    ✓ FOUND BUSINESS: {result_name} at {address[:40]}")
                    sys.stdout.flush()
                    return result_name
                print(f"    ✓ FOUND BUSINESS: {name} at {address[:40]}")
                sys.stdout.flush()
                return name

        print(f"    No results from Google for {address[:40]}")
        sys.stdout.flush()

    except Exception as e:
        print(f"    API Error for {address[:40]}: {str(e)[:50]}")
        sys.stdout.flush()

    return None

def lookup_business_at_address(address, use_cache=True):
    """Look up what business is at a given address using multiple methods

    All lookups are stored in business_mapping with source='google_api' or 'manual'
    """
    if not address:
        return None

    # Check mapping first (exact match) - includes both manual and API lookups
    mapped_name = get_mapping_name(address)
    if mapped_name is not None:
        # If we stored that no business was found, return None without lookup
        if mapped_name == "NO_BUSINESS_FOUND":
            return None
        return mapped_name

    # Check mapping with fuzzy matching (partial address match)
    # This allows mapping "10484 Beardslee Blvd, Bothell" to match "10484 Beardslee Blvd, Bothell WA 98011"
    normalized_address = normalize_address(address)
    for mapped_addr in business_mapping.keys():
        normalized_mapped = normalize_address(mapped_addr)
        # Check if the mapped address is contained in the actual address or vice versa
        if normalized_mapped in normalized_address or normalized_address in normalized_mapped:
            business_name = get_mapping_name(mapped_addr)
            if business_name and business_name != "NO_BUSINESS_FOUND":
                print(f"    ✓ FUZZY MATCHED: {business_name} (mapping: {mapped_addr[:30]})")
                sys.stdout.flush()
                return business_name

    business_name = None

    # Try Google Places API first (most accurate)
    if gmaps_client:
        business_name = lookup_business_google_places(address)
        if business_name:
            # Store in unified mapping with source
            set_mapping_entry(address, business_name, source="google_api")
            return business_name
        # If no business found, check for fuzzy matches before storing as NO_BUSINESS_FOUND
        else:
            # Check for similar addresses in mapping with business names
            fuzzy_matches = []
            for mapped_addr in business_mapping.keys():
                mapped_name = get_mapping_name(mapped_addr)
                # Skip NO_BUSINESS_FOUND entries
                if not mapped_name or mapped_name == "NO_BUSINESS_FOUND":
                    continue

                normalized_mapped = normalize_address(mapped_addr)
                # Check if the mapped address is contained in the actual address or vice versa
                if normalized_mapped in normalized_address or normalized_address in normalized_mapped:
                    # Additional check: they should share significant parts
                    address_parts = set(normalized_address.split())
                    mapped_parts = set(normalized_mapped.split())
                    common_parts = address_parts & mapped_parts

                    # If they share at least 3 parts, it's a potential match
                    if len(common_parts) >= 3:
                        fuzzy_matches.append({
                            'address': mapped_addr,
                            'business_name': mapped_name,
                            'common_parts': len(common_parts)
                        })

            # If we found fuzzy matches, ask the user
            if fuzzy_matches:
                # Sort by number of common parts (best match first)
                fuzzy_matches.sort(key=lambda x: x['common_parts'], reverse=True)
                best_match = fuzzy_matches[0]

                print()
                print(f"    ⚠ No business found at: {address}")
                print(f"    ✓ Found similar address: {best_match['address']}")
                print(f"      Business name: {best_match['business_name']}")
                print()

                try:
                    response = input(f"    Use this business name? (Y/n): ").strip().lower()
                    if response == '' or response == 'y' or response == 'yes':
                        print(f"    ✓ Using: {best_match['business_name']}")
                        sys.stdout.flush()
                        # Store this address with the matched business name
                        set_mapping_entry(address, best_match['business_name'], source="google_api")
                        return best_match['business_name']
                    else:
                        print(f"    Skipping fuzzy match")
                        sys.stdout.flush()
                except (EOFError, KeyboardInterrupt):
                    print()
                    print(f"    Skipping fuzzy match")
                    sys.stdout.flush()

            # Store the fact that no business was found
            set_mapping_entry(address, "NO_BUSINESS_FOUND", source="google_api")
            return None

    # Fall back to OpenStreetMap if Google didn't find anything
    if not GEOPY_AVAILABLE:
        return None

    try:
        # Initialize geocoder with a user agent
        geolocator = Nominatim(user_agent="dewart_mileage_tracker", timeout=10)

        # Add a small delay to respect rate limits (1 request per second)
        time.sleep(1.1)

        # Method 1: Try geocoding with addressdetails to get POI information
        location = geolocator.geocode(address, addressdetails=True, exactly_one=True)

        business_name = None

        if location and hasattr(location, 'raw'):
            raw = location.raw

            # Check for amenity, shop, or other business indicators
            if 'address' in raw:
                addr_details = raw['address']

                # Look for business-related fields
                for field in ['amenity', 'shop', 'tourism', 'leisure', 'building', 'office']:
                    if field in addr_details and addr_details[field]:
                        business_name = addr_details[field].replace('_', ' ').title()
                        break

                # Check for actual business name
                for field in ['name', 'brand', 'operator']:
                    if field in addr_details and addr_details[field]:
                        business_name = addr_details[field]
                        break

            # Method 2: Check the raw data for business information
            if not business_name:
                # Check class and type
                osm_class = raw.get('class', '')
                osm_type = raw.get('type', '')

                if osm_class == 'amenity':
                    if osm_type in ['fuel', 'charging_station']:
                        business_name = "Gas Station / Fuel"
                    elif osm_type in ['restaurant', 'cafe', 'fast_food']:
                        business_name = osm_type.replace('_', ' ').title()
                    elif osm_type in ['bank', 'post_office']:
                        business_name = osm_type.replace('_', ' ').title()
                elif osm_class == 'shop':
                    business_name = f"{osm_type.replace('_', ' ').title()} Shop"
                elif osm_class == 'office':
                    business_name = "Office Building"

            # Method 3: Try reverse geocoding at the same location
            if not business_name and location.latitude and location.longitude:
                time.sleep(1.1)
                reverse_results = geolocator.reverse(
                    (location.latitude, location.longitude),
                    exactly_one=False,
                    addressdetails=True,
                    zoom=18  # Building level zoom
                )

                if reverse_results:
                    for result in reverse_results[:3]:  # Check top 3 results
                        if hasattr(result, 'raw') and 'address' in result.raw:
                            addr = result.raw['address']
                            # Look for named places
                            if 'name' in result.raw and result.raw['name']:
                                name = result.raw['name']
                                # Avoid street names
                                if 'street' not in name.lower() and 'avenue' not in name.lower():
                                    business_name = name
                                    break

            # Store the result if found
            if business_name and business_name not in ['', 'yes']:
                set_mapping_entry(address, business_name, source="osm_api")
                return business_name
            else:
                # OpenStreetMap also failed - check for fuzzy matches before storing as not found
                fuzzy_matches = []
                for mapped_addr in business_mapping.keys():
                    mapped_name = get_mapping_name(mapped_addr)
                    # Skip NO_BUSINESS_FOUND entries
                    if not mapped_name or mapped_name == "NO_BUSINESS_FOUND":
                        continue

                    normalized_mapped = normalize_address(mapped_addr)
                    # Check if the mapped address is contained in the actual address or vice versa
                    if normalized_mapped in normalized_address or normalized_address in normalized_mapped:
                        # Additional check: they should share significant parts
                        address_parts = set(normalized_address.split())
                        mapped_parts = set(normalized_mapped.split())
                        common_parts = address_parts & mapped_parts

                        # If they share at least 3 parts, it's a potential match
                        if len(common_parts) >= 3:
                            fuzzy_matches.append({
                                'address': mapped_addr,
                                'business_name': mapped_name,
                                'common_parts': len(common_parts)
                            })

                # If we found fuzzy matches, ask the user
                if fuzzy_matches:
                    # Sort by number of common parts (best match first)
                    fuzzy_matches.sort(key=lambda x: x['common_parts'], reverse=True)
                    best_match = fuzzy_matches[0]

                    print()
                    print(f"    ⚠ No business found at: {address}")
                    print(f"    ✓ Found similar address: {best_match['address']}")
                    print(f"      Business name: {best_match['business_name']}")
                    print()

                    try:
                        response = input(f"    Use this business name? (Y/n): ").strip().lower()
                        if response == '' or response == 'y' or response == 'yes':
                            print(f"    ✓ Using: {best_match['business_name']}")
                            sys.stdout.flush()
                            # Store this address with the matched business name
                            set_mapping_entry(address, best_match['business_name'], source="osm_api")
                            return best_match['business_name']
                        else:
                            print(f"    Skipping fuzzy match")
                            sys.stdout.flush()
                    except (EOFError, KeyboardInterrupt):
                        print()
                        print(f"    Skipping fuzzy match")
                        sys.stdout.flush()

                # Store that no business was found (OpenStreetMap also failed)
                set_mapping_entry(address, "NO_BUSINESS_FOUND", source="osm_api")
                return None

    except (GeocoderTimedOut, GeocoderServiceError) as e:
        # Don't print errors for every lookup - they're cached anyway
        # Check for fuzzy matches before storing as failed
        fuzzy_matches = []
        for mapped_addr in business_mapping.keys():
            mapped_name = get_mapping_name(mapped_addr)
            if not mapped_name or mapped_name == "NO_BUSINESS_FOUND":
                continue

            normalized_mapped = normalize_address(mapped_addr)
            if normalized_mapped in normalized_address or normalized_address in normalized_mapped:
                address_parts = set(normalized_address.split())
                mapped_parts = set(normalized_mapped.split())
                common_parts = address_parts & mapped_parts

                if len(common_parts) >= 3:
                    fuzzy_matches.append({
                        'address': mapped_addr,
                        'business_name': mapped_name,
                        'common_parts': len(common_parts)
                    })

        if fuzzy_matches:
            fuzzy_matches.sort(key=lambda x: x['common_parts'], reverse=True)
            best_match = fuzzy_matches[0]

            print()
            print(f"    ⚠ Lookup timed out for: {address}")
            print(f"    ✓ Found similar address: {best_match['address']}")
            print(f"      Business name: {best_match['business_name']}")
            print()

            try:
                response = input(f"    Use this business name? (Y/n): ").strip().lower()
                if response == '' or response == 'y' or response == 'yes':
                    print(f"    ✓ Using: {best_match['business_name']}")
                    sys.stdout.flush()
                    set_mapping_entry(address, best_match['business_name'], source="osm_api")
                    return best_match['business_name']
                else:
                    print(f"    Skipping fuzzy match")
                    sys.stdout.flush()
            except (EOFError, KeyboardInterrupt):
                print()
                print(f"    Skipping fuzzy match")
                sys.stdout.flush()

        set_mapping_entry(address, "NO_BUSINESS_FOUND", source="osm_api")
        pass
    except Exception as e:
        # Silently fail for individual lookups
        # Still store the failed lookup
        set_mapping_entry(address, "NO_BUSINESS_FOUND", source="osm_api")
        pass

    return None

def get_business_name(address, lookup=False):
    """Extract potential business name from address"""
    if not address:
        return "Unknown"

    # Try online lookup FIRST if enabled
    if lookup and GEOPY_AVAILABLE:
        looked_up = lookup_business_at_address(address)
        if looked_up:
            return looked_up

    # Check for known business types from keywords in address
    norm = normalize_address(address)
    if any(gas in norm for gas in ["gas", "fuel", "shell", "chevron", "76", "arco", "bp", "exxon", "mobil"]):
        return "Gas Station"
    if any(store in norm for store in ["costco", "safeway", "qfc", "fred meyer", "walmart", "target"]):
        return "Store/Shopping"
    if "pioneer" in norm or "state route" in norm:
        return "Business Location"

    # Return empty string if no business name found (per user request)
    return ""

def categorize_trip(trip, prev_trip=None, next_trip=None, enable_lookup=False):
    """Categorize a trip as commute, business, or personal"""
    start_addr = trip['start_address']
    end_addr = trip['end_address']
    distance = trip['distance']
    timestamp = trip['started']

    # Check if we have a saved category in the business mapping (highest priority)
    saved_category = get_mapping_category(end_addr)
    if saved_category:
        # Use saved category with business name lookup
        business_name = get_business_name(end_addr, lookup=enable_lookup)
        return saved_category.lower(), business_name

    # Gas stations are always business
    if is_business_location(start_addr) or is_business_location(end_addr):
        return 'business', get_business_name(start_addr or end_addr, lookup=enable_lookup)

    # Home to work or work to home = commute
    if is_home_address(start_addr) and is_work_address(end_addr):
        return 'commute', 'Office'
    if is_work_address(start_addr) and is_home_address(end_addr):
        return 'commute', 'Home'

    # Trips to/from Whidbey - all personal per client request
    if is_whidbey_area(start_addr) or is_whidbey_area(end_addr):
        return 'personal', 'Whidbey Personal Trip'

    # Trips over threshold during weekday = business
    if distance >= BUSINESS_DISTANCE_THRESHOLD:
        day_of_week = timestamp.weekday()
        if day_of_week < 5:  # Weekday
            # Check if destination is home or work FIRST
            if is_home_address(end_addr):
                return 'business', "Home"
            elif is_work_address(end_addr):
                return 'business', "Office"
            else:
                business_name = get_business_name(end_addr, lookup=enable_lookup)
                return 'business', business_name

    # Weekend trips (Friday evening to Monday morning)
    day_of_week = timestamp.weekday()
    hour = timestamp.hour
    if day_of_week == 4 and hour >= 17:  # Friday after 5pm
        return 'personal', 'Weekend Travel'
    if day_of_week in [5, 6]:  # Saturday, Sunday
        return 'personal', 'Weekend Travel'
    if day_of_week == 0 and hour < 7:  # Monday before 7am
        return 'personal', 'Weekend Travel'

    # Short trips in Bothell/Kenmore area during weekday
    if is_bothell_area(start_addr) and is_bothell_area(end_addr):
        if distance < BUSINESS_DISTANCE_THRESHOLD:
            day_of_week = timestamp.weekday()
            if day_of_week < 5:
                # Check if destination is home or work FIRST
                if is_home_address(end_addr):
                    return 'business', "Home"
                elif is_work_address(end_addr):
                    return 'business', "Office"
                else:
                    # For local business trips, try to get actual business name
                    business_name = get_business_name(end_addr, lookup=enable_lookup)
                    return 'business', business_name
            else:
                return 'personal', 'Local Personal'

    # Default to personal
    return 'personal', 'Other Personal'

def get_week_key(date):
    """Get week identifier (year-week)"""
    # Week starts on Monday
    week_start = date - timedelta(days=date.weekday())
    return week_start.strftime('%Y-%m-%d')

def parse_distance(dist_str):
    """Parse distance string to float"""
    if not dist_str:
        return 0.0
    try:
        # Remove any non-numeric characters except decimal point
        cleaned = re.sub(r'[^\d.]', '', str(dist_str))
        return float(cleaned) if cleaned else 0.0
    except:
        return 0.0

def merge_short_stops(trips, max_gap_minutes=3, max_stop_distance=0.2):
    """
    Merge trips that are likely false stops (red lights, traffic, etc.)

    Heuristics:
    - Gap between trips is very short (< max_gap_minutes, default 3 min)
    - The "stop" trip has a very short distance (< max_stop_distance, default 0.2 miles)
    - The end address of first trip matches (or is very close to) start of next trip

    When merged:
    - Keep start time and start address from first trip
    - Keep end time and end address from last trip
    - Sum the distances
    - Store merged trip info for transparency

    Returns: (merged_trips, merge_count)
    """
    if not trips or len(trips) < 2:
        return trips, 0

    # Sort by start time to ensure proper ordering
    sorted_trips = sorted(trips, key=lambda x: x['started'])

    merged = []
    merge_count = 0
    i = 0

    while i < len(sorted_trips):
        current = sorted_trips[i].copy()  # Copy to avoid modifying original

        # Look ahead to see if we should merge with next trip(s)
        while i + 1 < len(sorted_trips):
            next_trip = sorted_trips[i + 1]

            # Parse stop time from current trip
            current_stop = None
            if 'stopped' in current and current['stopped']:
                if isinstance(current['stopped'], datetime):
                    current_stop = current['stopped']
                else:
                    try:
                        current_stop = datetime.strptime(str(current['stopped']).strip(), '%Y-%m-%d %H:%M')
                    except:
                        pass

            next_start = next_trip['started']

            if current_stop is None:
                # Can't determine gap without stop time
                break

            # Calculate gap between trips
            gap = next_start - current_stop
            gap_minutes = gap.total_seconds() / 60

            # Check if this looks like a false stop
            # Either the current trip is very short OR the next trip is very short
            # AND the gap is small
            current_distance = current.get('distance', 0)
            next_distance = next_trip.get('distance', 0)

            should_merge = False

            if gap_minutes <= max_gap_minutes and gap_minutes >= 0:
                # Short gap - check distances
                if current_distance <= max_stop_distance or next_distance <= max_stop_distance:
                    # At least one trip is very short - likely a false stop
                    should_merge = True

            if not should_merge:
                break

            # Merge the trips
            merge_count += 1

            # Track what was merged for transparency
            if 'merged_from' not in current:
                current['merged_from'] = [{
                    'started': current['started'].strftime('%Y-%m-%d %H:%M'),
                    'stopped': str(current.get('stopped', '')),
                    'distance': current_distance,
                    'start_address': current.get('start_address', ''),
                    'end_address': current.get('end_address', '')
                }]

            current['merged_from'].append({
                'started': next_trip['started'].strftime('%Y-%m-%d %H:%M'),
                'stopped': str(next_trip.get('stopped', '')),
                'distance': next_distance,
                'start_address': next_trip.get('start_address', ''),
                'end_address': next_trip.get('end_address', '')
            })

            # Update current trip with merged data
            # Keep start from current, take end from next
            current['stopped'] = next_trip.get('stopped', current.get('stopped'))
            current['end_address'] = next_trip.get('end_address', current.get('end_address'))
            current['end_odometer'] = next_trip.get('end_odometer', current.get('end_odometer'))
            current['distance'] = current_distance + next_distance

            # Mark as merged for display purposes
            current['is_merged'] = True
            current['merge_count'] = len(current['merged_from'])

            i += 1  # Skip the merged trip

        merged.append(current)
        i += 1

    return merged, merge_count


def flag_micro_trips(trips, max_distance=0.15):
    """
    Flag micro-trips that are suspiciously short and likely GPS drift or parking adjustments.

    Criteria for micro-trip:
    - Distance is very small (< max_distance, default 0.15 miles)
    - Start and end addresses are on the same street or very close

    These are flagged but NOT removed - the user decides what to do with them.

    Returns: (trips_with_flags, micro_count)
    """
    micro_count = 0

    for trip in trips:
        distance = trip.get('distance', 0)
        start_addr = trip.get('start_address', '').lower()
        end_addr = trip.get('end_address', '').lower()

        is_micro = False

        # Check if distance is very small
        if distance <= max_distance:
            # Check if addresses are similar (same street)
            # Extract street name from addresses
            start_parts = start_addr.replace(',', ' ').split()
            end_parts = end_addr.replace(',', ' ').split()

            # Look for common street identifiers
            street_types = ['st', 'ave', 'rd', 'dr', 'ln', 'way', 'blvd', 'ct', 'pl', 'circle',
                           'street', 'avenue', 'road', 'drive', 'lane', 'boulevard', 'court', 'place']

            def extract_street(parts):
                """Extract street name from address parts"""
                for i, part in enumerate(parts):
                    if part.lower().rstrip('.') in street_types and i > 0:
                        # Return the part before the street type and the type
                        return ' '.join(parts[max(0,i-2):i+1]).lower()
                return ' '.join(parts[:3]).lower() if len(parts) >= 3 else ' '.join(parts).lower()

            start_street = extract_street(start_parts)
            end_street = extract_street(end_parts)

            # If same street or very similar, it's likely a micro-trip
            if start_street == end_street:
                is_micro = True
            elif distance <= 0.1:
                # Very short distance - likely micro-trip regardless of street match
                is_micro = True

        if is_micro:
            trip['is_micro_trip'] = True
            trip['micro_reason'] = f"Very short trip ({distance:.2f} mi)"
            if start_addr and end_addr:
                # Check if essentially same location
                if distance <= 0.05:
                    trip['micro_reason'] = f"GPS drift or parking adjustment ({distance:.2f} mi)"
            micro_count += 1
        else:
            trip['is_micro_trip'] = False

    return trips, micro_count


def read_trips(input_file):
    """Read trips from CSV or XLSX file"""
    trips = []

    # Check file extension
    file_ext = os.path.splitext(input_file)[1].lower()

    if file_ext == '.xlsx':
        # Read from Excel file
        return read_trips_from_xlsx(input_file)
    else:
        # Read from CSV file (default)
        return read_trips_from_csv(input_file)

def read_trips_from_xlsx(xlsx_file):
    """Read trips from Excel file"""
    trips = []

    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl is required to read Excel files", file=sys.stderr)
        return trips

    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file, read_only=False, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val else '')

        # Map column names to indices (1-based for openpyxl)
        col_map = {h: i + 1 for i, h in enumerate(headers)}

        print(f"  Found {len(headers)} columns, {ws.max_row - 1} data rows")
        sys.stdout.flush()

        for row_num in range(2, ws.max_row + 1):
            try:
                # Get Started value
                started_col = col_map.get('Started', 2)
                started_val = ws.cell(row=row_num, column=started_col).value
                if not started_val:
                    continue

                # Parse date - could be string or datetime object
                if isinstance(started_val, datetime):
                    started = started_val
                else:
                    started = datetime.strptime(str(started_val).strip(), '%Y-%m-%d %H:%M')

                # Get distance
                dist_col = col_map.get('Distance (miles)', 9)
                dist_val = ws.cell(row=row_num, column=dist_col).value
                distance = parse_distance(str(dist_val) if dist_val else '0')

                # Helper function to get cell value safely
                def get_val(col_name, default_col):
                    col = col_map.get(col_name, default_col)
                    val = ws.cell(row=row_num, column=col).value
                    return str(val).strip() if val else ''

                trip = {
                    'started': started,
                    'start_odometer': get_val('Start odometer (miles)', 3),
                    'start_address': get_val('Start address', 4),
                    'stopped': get_val('Stopped', 5),
                    'end_odometer': get_val('End odometer (miles)', 6),
                    'end_address': get_val('End address', 7),
                    'distance': distance,
                    'category': get_val('Category', 1),
                    'user_notes': get_val('User Notes', 14)
                }

                trips.append(trip)
            except Exception as e:
                print(f"Warning: Skipping row {row_num} due to parsing error: {e}", file=sys.stderr)
                continue

        wb.close()
    except Exception as e:
        print(f"Error reading Excel file: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()

    # Sort by date (oldest first)
    trips.sort(key=lambda x: x['started'])
    return trips

def read_trips_from_csv(csv_file):
    """Read trips from CSV file"""
    trips = []

    try:
        with open(csv_file, 'r', encoding='utf-16-le') as f:
            content = f.read()
            # Remove BOM if present
            if content.startswith('\ufeff'):
                content = content[1:]

            # Parse CSV from string
            from io import StringIO
            reader = csv.DictReader(StringIO(content), delimiter=';')

            for row in reader:
                try:
                    # Skip empty rows
                    if not row.get('Started'):
                        continue

                    # Parse the row
                    started = datetime.strptime(row['Started'].strip(), '%Y-%m-%d %H:%M')
                    distance = parse_distance(row['Distance (miles)'])

                    trip = {
                        'started': started,
                        'start_odometer': row['Start odometer (miles)'].strip(),
                        'start_address': row['Start address'].strip(),
                        'stopped': row['Stopped'].strip(),
                        'end_odometer': row['End odometer (miles)'].strip(),
                        'end_address': row['End address'].strip(),
                        'distance': distance,
                        'category': row['Category'].strip(),
                        'user_notes': row.get('User Notes', '').strip()
                    }

                    trips.append(trip)
                except Exception as e:
                    print(f"Warning: Skipping row due to parsing error: {e}", file=sys.stderr)
                    continue
    except UnicodeError:
        # Try UTF-16 BE
        with open(csv_file, 'r', encoding='utf-16-be') as f:
            content = f.read()
            if content.startswith('\ufeff'):
                content = content[1:]

            from io import StringIO
            reader = csv.DictReader(StringIO(content), delimiter=';')

            for row in reader:
                try:
                    if not row.get('Started'):
                        continue

                    started = datetime.strptime(row['Started'].strip(), '%Y-%m-%d %H:%M')
                    distance = parse_distance(row['Distance (miles)'])

                    trip = {
                        'started': started,
                        'start_odometer': row['Start odometer (miles)'].strip(),
                        'start_address': row['Start address'].strip(),
                        'stopped': row['Stopped'].strip(),
                        'end_odometer': row['End odometer (miles)'].strip(),
                        'end_address': row['End address'].strip(),
                        'distance': distance,
                        'category': row['Category'].strip(),
                        'user_notes': row.get('User Notes', '').strip()
                    }

                    trips.append(trip)
                except Exception as e:
                    print(f"Warning: Skipping row due to parsing error: {e}", file=sys.stderr)
                    continue

    # Sort by date (oldest first)
    trips.sort(key=lambda x: x['started'])

    return trips

def export_weekly_summary_csv(weekly_stats, filename="weekly_summary.csv"):
    """Export weekly summary to CSV file"""
    try:
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Week Starting', 'Commute Miles', 'Business Miles', 'Personal Miles',
                           'Total Miles', 'Portland Round Trip Miles', 'Spokane Round Trip Miles',
                           'Weekend Miles (Fri 5pm-Mon 6am)'])

            total_commute = 0.0
            total_business = 0.0
            total_personal = 0.0
            total_all = 0.0
            total_portland = 0.0
            total_spokane = 0.0
            total_weekend = 0.0

            for week in sorted(weekly_stats.keys()):
                stats = weekly_stats[week]
                commute = stats['commute']
                business = stats['business']
                personal = stats['personal']
                total = stats['total']
                portland = stats.get('portland_miles', 0.0)
                spokane = stats.get('spokane_miles', 0.0)
                weekend = stats.get('weekend_miles', 0.0)

                writer.writerow([week, f"{commute:.1f}", f"{business:.1f}", f"{personal:.1f}",
                               f"{total:.1f}", f"{portland:.1f}", f"{spokane:.1f}", f"{weekend:.1f}"])

                total_commute += commute
                total_business += business
                total_personal += personal
                total_all += total
                total_portland += portland
                total_spokane += spokane
                total_weekend += weekend

            writer.writerow(['TOTAL', f"{total_commute:.1f}", f"{total_business:.1f}",
                           f"{total_personal:.1f}", f"{total_all:.1f}", f"{total_portland:.1f}",
                           f"{total_spokane:.1f}", f"{total_weekend:.1f}"])

        return True
    except Exception as e:
        print(f"Error exporting weekly summary: {e}", file=sys.stderr)
        return False

def export_detailed_trips_csv(categorized_trips, filename="detailed_trips.csv"):
    """Export detailed trip log to CSV file"""
    try:
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Date', 'Time', 'Day of Week', 'Category', 'Distance (miles)',
                           'Start Address', 'End Address', 'Business Name',
                           'Start Odometer', 'End Odometer', 'Week Starting'])

            for trip in categorized_trips:
                date_str = trip['started'].strftime('%Y-%m-%d')
                time_str = trip['started'].strftime('%H:%M')
                day_of_week = trip['started'].strftime('%A')
                category = trip['auto_category'].title()
                distance = f"{trip['distance']:.1f}"
                start_addr = trip['start_address']
                end_addr = trip['end_address']
                business_name = trip['business_name'] if trip['auto_category'] in ['business', 'commute'] else ''
                start_odo = trip['start_odometer']
                end_odo = trip['end_odometer']
                week = get_week_key(trip['started'])

                writer.writerow([date_str, time_str, day_of_week, category, distance,
                               start_addr, end_addr, business_name, start_odo, end_odo, week])

        return True
    except Exception as e:
        print(f"Error exporting detailed trips: {e}", file=sys.stderr)
        return False

def export_summary_csv(total_commute, total_business, total_personal, total_all, filename="summary.csv"):
    """Export overall summary to CSV file"""
    try:
        actual_business = total_business
        actual_personal = total_personal + total_commute

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Category', 'Miles', 'Percentage'])
            writer.writerow(['Total Miles Driven', f"{total_all:.1f}", '100.0%'])
            writer.writerow(['Business Miles', f"{actual_business:.1f}", f"{actual_business/total_all*100:.1f}%"])
            writer.writerow(['Personal Miles', f"{actual_personal:.1f}", f"{actual_personal/total_all*100:.1f}%"])
            writer.writerow(['  - Commute', f"{total_commute:.1f}", f"{total_commute/total_all*100:.1f}%"])
            writer.writerow(['  - Other Personal', f"{total_personal:.1f}", f"{total_personal/total_all*100:.1f}%"])

        return True
    except Exception as e:
        print(f"Error exporting summary: {e}", file=sys.stderr)
        return False

def reconstruct_full_trips(categorized_trips):
    """Reconstruct full trips from individual segments"""
    if not categorized_trips:
        return []

    # Sort trips by start time
    sorted_trips = sorted(categorized_trips, key=lambda x: x['started'])

    full_trips = []
    current_trip = None

    for trip in sorted_trips:
        # Parse end time from stopped field
        try:
            end_time = datetime.strptime(trip['stopped'], '%Y-%m-%d %H:%M')
        except:
            # Estimate end time as start + 30 minutes if parsing fails
            end_time = trip['started'] + timedelta(minutes=30)

        # Start a new full trip if:
        # 1. This is the first trip
        # 2. Time gap > 30 minutes from last segment's end
        # 3. The start address doesn't match the previous end address

        if current_trip is None:
            # Start first trip
            current_trip = {
                'segments': [trip],
                'start_time': trip['started'],
                'end_time': end_time,
                'start_address': trip['start_address'],
                'end_address': trip['end_address'],
                'total_distance': trip['distance'],
                'stops': [],
                'primary_category': trip['auto_category'],
                'business_stops': []
            }
        else:
            # Check if this segment continues the current trip
            time_gap = (trip['started'] - current_trip['end_time']).total_seconds() / 60
            addresses_match = normalize_address(trip['start_address']) == normalize_address(current_trip['end_address'])

            if time_gap <= 30 and addresses_match:
                # Continue current trip
                current_trip['segments'].append(trip)
                current_trip['end_time'] = end_time
                current_trip['end_address'] = trip['end_address']
                current_trip['total_distance'] += trip['distance']

                # Add stop (the intermediate location)
                current_trip['stops'].append({
                    'address': trip['start_address'],
                    'time': trip['started']
                })

                # Track business stops
                if trip['auto_category'] == 'business' and trip['business_name']:
                    current_trip['business_stops'].append({
                        'business_name': trip['business_name'],
                        'address': trip['end_address'],
                        'distance': trip['distance']
                    })

                # Update primary category (business takes priority)
                if trip['auto_category'] == 'business':
                    current_trip['primary_category'] = 'business'
            else:
                # Save current trip and start new one
                full_trips.append(current_trip)
                current_trip = {
                    'segments': [trip],
                    'start_time': trip['started'],
                    'end_time': end_time,
                    'start_address': trip['start_address'],
                    'end_address': trip['end_address'],
                    'total_distance': trip['distance'],
                    'stops': [],
                    'primary_category': trip['auto_category'],
                    'business_stops': []
                }

    # Don't forget the last trip
    if current_trip:
        full_trips.append(current_trip)

    return full_trips

def export_reconstructed_trips_csv(categorized_trips, filename="reconstructed_trips.csv"):
    """Export reconstructed full trips"""
    try:
        full_trips = reconstruct_full_trips(categorized_trips)

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Date', 'Start Time', 'End Time', 'Duration (min)', 'Category',
                           'Start Address', 'End Address', 'Stops', 'Business Stops',
                           'Total Distance (miles)', 'Segment Count'])

            for trip in full_trips:
                date_str = trip['start_time'].strftime('%Y-%m-%d')
                start_time = trip['start_time'].strftime('%H:%M')
                end_time = trip['end_time'].strftime('%H:%M')
                duration = (trip['end_time'] - trip['start_time']).total_seconds() / 60
                category = trip['primary_category'].title()

                # Format stops
                stops_str = ' -> '.join([s['address'] for s in trip['stops']]) if trip['stops'] else 'Direct'

                # Format business stops
                business_stops_str = '; '.join([f"{b['business_name']} ({b['distance']:.1f}mi)"
                                               for b in trip['business_stops']]) if trip['business_stops'] else ''

                writer.writerow([date_str, start_time, end_time, f"{duration:.0f}",
                               category, trip['start_address'], trip['end_address'],
                               stops_str, business_stops_str,
                               f"{trip['total_distance']:.1f}", len(trip['segments'])])

        return True
    except Exception as e:
        print(f"Error exporting reconstructed trips: {e}", file=sys.stderr)
        return False

def export_address_business_correlation_csv(categorized_trips, filename="address_business_correlation.csv"):
    """Export address to business name correlation with total mileage"""
    try:
        # Collect address data with business names and mileage
        address_data = defaultdict(lambda: {'business_name': '', 'total_miles': 0.0, 'trip_count': 0})

        for trip in categorized_trips:
            # Only track business trips with business names
            if trip['auto_category'] == 'business' and trip['business_name']:
                end_addr = trip['end_address']
                if end_addr:
                    # Use the business name from the trip
                    if not address_data[end_addr]['business_name']:
                        address_data[end_addr]['business_name'] = trip['business_name']
                    address_data[end_addr]['total_miles'] += trip['distance']
                    address_data[end_addr]['trip_count'] += 1

        # Sort by total miles (descending)
        sorted_addresses = sorted(address_data.items(),
                                 key=lambda x: x[1]['total_miles'],
                                 reverse=True)

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Address', 'Business Name', 'Total Miles', 'Trip Count'])

            for address, data in sorted_addresses:
                writer.writerow([address, data['business_name'],
                               f"{data['total_miles']:.1f}", data['trip_count']])

        return True
    except Exception as e:
        print(f"Error exporting address-business correlation: {e}", file=sys.stderr)
        return False

def export_comprehensive_report_csv(categorized_trips, weekly_stats, filename="comprehensive_report.csv"):
    """Export comprehensive report with all key information in one place"""
    try:
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Section 1: Overall Summary
            writer.writerow(['OVERALL SUMMARY'])
            writer.writerow([])

            total_commute = sum(stats['commute'] for stats in weekly_stats.values())
            total_business = sum(stats['business'] for stats in weekly_stats.values())
            total_personal = sum(stats['personal'] for stats in weekly_stats.values())
            total_all = sum(stats['total'] for stats in weekly_stats.values())
            total_portland = sum(stats.get('portland_miles', 0.0) for stats in weekly_stats.values())
            total_spokane = sum(stats.get('spokane_miles', 0.0) for stats in weekly_stats.values())
            total_weekend = sum(stats.get('weekend_miles', 0.0) for stats in weekly_stats.values())

            writer.writerow(['Metric', 'Miles', 'Percentage'])
            writer.writerow(['Total Miles Driven', f"{total_all:.1f}", '100%'])
            writer.writerow(['Business Miles', f"{total_business:.1f}", f"{total_business/total_all*100:.1f}%"])
            writer.writerow(['Personal Miles (incl. Commute)', f"{total_personal + total_commute:.1f}",
                           f"{(total_personal + total_commute)/total_all*100:.1f}%"])
            writer.writerow(['  - Commute', f"{total_commute:.1f}", f"{total_commute/total_all*100:.1f}%"])
            writer.writerow(['  - Other Personal', f"{total_personal:.1f}", f"{total_personal/total_all*100:.1f}%"])
            writer.writerow([])

            writer.writerow(['Special Tracking', 'Miles'])
            writer.writerow(['Portland Round Trip Miles', f"{total_portland:.1f}"])
            writer.writerow(['Spokane Round Trip Miles', f"{total_spokane:.1f}"])
            writer.writerow(['Weekend Miles (Fri 5pm-Mon 6am)', f"{total_weekend:.1f}"])
            writer.writerow([])
            writer.writerow([])

            # Section 2: Weekly Breakdown
            writer.writerow(['WEEKLY BREAKDOWN'])
            writer.writerow([])
            writer.writerow(['Week Starting', 'Commute', 'Business', 'Personal', 'Total',
                           'Portland RT', 'Spokane RT', 'Weekend'])

            for week in sorted(weekly_stats.keys()):
                stats = weekly_stats[week]
                writer.writerow([week, f"{stats['commute']:.1f}", f"{stats['business']:.1f}",
                               f"{stats['personal']:.1f}", f"{stats['total']:.1f}",
                               f"{stats.get('portland_miles', 0.0):.1f}",
                               f"{stats.get('spokane_miles', 0.0):.1f}",
                               f"{stats.get('weekend_miles', 0.0):.1f}"])

            writer.writerow(['TOTAL', f"{total_commute:.1f}", f"{total_business:.1f}",
                           f"{total_personal:.1f}", f"{total_all:.1f}",
                           f"{total_portland:.1f}", f"{total_spokane:.1f}", f"{total_weekend:.1f}"])
            writer.writerow([])
            writer.writerow([])

            # Section 3: Top Business Destinations
            writer.writerow(['TOP BUSINESS DESTINATIONS'])
            writer.writerow([])

            # Collect business destination data
            business_destinations = defaultdict(lambda: {'miles': 0.0, 'trips': 0})
            for trip in categorized_trips:
                if trip['auto_category'] == 'business' and trip['business_name']:
                    key = f"{trip['business_name']} - {trip['end_address']}"
                    business_destinations[key]['miles'] += trip['distance']
                    business_destinations[key]['trips'] += 1

            # Sort by miles
            sorted_destinations = sorted(business_destinations.items(),
                                        key=lambda x: x[1]['miles'],
                                        reverse=True)

            writer.writerow(['Business Destination', 'Total Miles', 'Trip Count'])
            for dest, data in sorted_destinations[:20]:  # Top 20
                writer.writerow([dest, f"{data['miles']:.1f}", data['trips']])
            writer.writerow([])
            writer.writerow([])

            # Section 4: Reconstructed Trip Summary
            writer.writerow(['RECONSTRUCTED TRIPS SUMMARY'])
            writer.writerow([])

            full_trips = reconstruct_full_trips(categorized_trips)

            # Count trips by type
            trip_counts = defaultdict(int)
            trip_miles = defaultdict(float)
            for ft in full_trips:
                trip_counts[ft['primary_category']] += 1
                trip_miles[ft['primary_category']] += ft['total_distance']

            writer.writerow(['Category', 'Full Trips', 'Total Miles', 'Avg Miles/Trip'])
            for category in ['business', 'commute', 'personal']:
                count = trip_counts[category]
                miles = trip_miles[category]
                avg = miles / count if count > 0 else 0
                writer.writerow([category.title(), count, f"{miles:.1f}", f"{avg:.1f}"])

            writer.writerow([])
            writer.writerow(['Recent Reconstructed Trips (Last 30)'])
            writer.writerow(['Date', 'Start', 'End', 'Duration', 'Category', 'Start -> End',
                           'Stops', 'Business Stops', 'Miles'])

            # Show last 30 reconstructed trips
            for trip in full_trips[-30:]:
                date_str = trip['start_time'].strftime('%Y-%m-%d')
                start_time = trip['start_time'].strftime('%H:%M')
                end_time = trip['end_time'].strftime('%H:%M')
                duration = f"{(trip['end_time'] - trip['start_time']).total_seconds() / 60:.0f}min"
                route = f"{trip['start_address'][:30]} -> {trip['end_address'][:30]}"
                stops = str(len(trip['stops']))
                business_stops = '; '.join([b['business_name'] for b in trip['business_stops']])

                writer.writerow([date_str, start_time, end_time, duration,
                               trip['primary_category'].title(), route, stops,
                               business_stops, f"{trip['total_distance']:.1f}"])

        return True
    except Exception as e:
        print(f"Error exporting comprehensive report: {e}", file=sys.stderr)
        return False

def export_route_summary_csv(categorized_trips, filename="route_summary.csv"):
    """Export route summary with individual trip details"""
    try:
        # Group trips by route (start -> end address pair)
        routes = defaultdict(list)

        for trip in categorized_trips:
            route_key = (trip['start_address'], trip['end_address'])
            routes[route_key].append(trip)

        # Sort routes by total distance (descending)
        sorted_routes = sorted(routes.items(),
                             key=lambda x: sum(t['distance'] for t in x[1]),
                             reverse=True)

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Start Address', 'End Address', 'Date', 'Time', 'Distance (miles)',
                           'Category', 'Business Name', 'Trip Count', 'Total Distance'])

            for (start_addr, end_addr), trips in sorted_routes:
                trip_count = len(trips)
                total_dist = sum(t['distance'] for t in trips)

                # Sort trips by date
                trips_sorted = sorted(trips, key=lambda x: x['started'])

                # Write summary row
                writer.writerow([start_addr, end_addr, '', '', '', '', '',
                               f'{trip_count} trips', f'{total_dist:.1f}'])

                # Write individual trip rows
                for trip in trips_sorted:
                    date_str = trip['started'].strftime('%Y-%m-%d')
                    time_str = trip['started'].strftime('%H:%M')
                    distance = f"{trip['distance']:.1f}"
                    category = trip['auto_category'].title()
                    business_name = trip['business_name'] if trip['auto_category'] in ['business', 'commute'] else ''

                    writer.writerow(['', '', date_str, time_str, distance, category, business_name, '', ''])

                # Add blank row between routes
                writer.writerow(['', '', '', '', '', '', '', '', ''])

        return True
    except Exception as e:
        print(f"Error exporting route summary: {e}", file=sys.stderr)
        return False

def analyze_mileage(csv_file, enable_lookup=False, start_date=None, end_date=None):
    """Main analysis function"""

    # Print startup message IMMEDIATELY
    print("=" * 80)
    print("MILEAGE ANALYZER - STARTING")
    print("=" * 80)
    print()
    sys.stdout.flush()  # Force output to appear immediately

    # Load business mapping and config if lookup is enabled
    if enable_lookup:
        print("Loading configuration and mapping files...")
        sys.stdout.flush()
        load_business_mapping()
        load_config()

        if gmaps_client:
            print("  [OK] Google Places API configured")
        else:
            print("  [OK] Will use OpenStreetMap for lookups")
        print()
        sys.stdout.flush()

    print(f"Reading CSV file: {csv_file}")
    print("Please wait, this may take a moment for large files...")
    sys.stdout.flush()
    trips = read_trips(csv_file)
    print(f"  [OK] Loaded {len(trips)} trips")
    print()
    sys.stdout.flush()

    if not trips:
        print("No trips found in CSV file!")
        return

    # Prompt for date range if not provided
    if start_date is None and end_date is None:
        # Find min/max dates in data
        min_date = min(t['started'] for t in trips).strftime('%Y-%m-%d')
        max_date = max(t['started'] for t in trips).strftime('%Y-%m-%d')

        print("-" * 80)
        print("DATE RANGE")
        print("-" * 80)
        print(f"Data available: {min_date} to {max_date}")
        print()
        print("Enter date range or press Enter for all data:")
        sys.stdout.flush()

        try:
            start_input = input("  Start date (YYYY-MM-DD) [all]: ").strip()
            if start_input:
                start_date = start_input
                end_input = input("  End date (YYYY-MM-DD) [all]: ").strip()
                if end_input:
                    end_date = end_input
        except (EOFError, KeyboardInterrupt):
            pass
        print()
        sys.stdout.flush()

    # Filter trips by date range
    if start_date or end_date:
        original_count = len(trips)

        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                trips = [t for t in trips if t['started'] >= start_dt]
            except ValueError:
                print(f"Warning: Invalid start date '{start_date}', ignoring")

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                trips = [t for t in trips if t['started'] <= end_dt]
            except ValueError:
                print(f"Warning: Invalid end date '{end_date}', ignoring")

        print(f"Filtered to {len(trips)} trips (from {original_count})")
        if start_date and end_date:
            print(f"Date range: {start_date} to {end_date}")
        elif start_date:
            print(f"From: {start_date}")
        elif end_date:
            print(f"Until: {end_date}")
        print()
        sys.stdout.flush()

        if not trips:
            print("No trips found in specified date range!")
            return

    print("=" * 80)
    print("MILEAGE ANALYSIS REPORT")
    print("D'Ewart Representatives, L.L.C.")
    print("=" * 80)
    print()

    # Categorize all trips
    categorized_trips = []

    if enable_lookup:
        lookup_method = "Google Places API" if gmaps_client else "OpenStreetMap"
        print(f"Categorizing and looking up businesses for {len(trips)} trips...")
        print(f"Using: {lookup_method}")
        print("This may take a while for the first run (results are cached)...\n")
        sys.stdout.flush()

    for i, trip in enumerate(trips):
        prev_trip = trips[i-1] if i > 0 else None
        next_trip = trips[i+1] if i < len(trips)-1 else None

        # Categorize trip and get business name (with lookup if enabled)
        category, business_name = categorize_trip(trip, prev_trip, next_trip, enable_lookup)

        trip['auto_category'] = category
        trip['business_name'] = business_name
        categorized_trips.append(trip)

        # Show progress for lookups
        if enable_lookup and category == 'business' and (i + 1) % 10 == 0:
            print(f"  Processed {i + 1}/{len(trips)} trips...")
            sys.stdout.flush()

    # Save the mapping if we did any lookups
    if enable_lookup:
        save_business_mapping()
        print("Business lookup complete!\n")
        sys.stdout.flush()

    # Calculate statistics by week
    weekly_stats = defaultdict(lambda: {
        'commute': 0.0,
        'business': 0.0,
        'personal': 0.0,
        'total': 0.0,
        'portland_miles': 0.0,
        'spokane_miles': 0.0,
        'weekend_miles': 0.0,
        'trips': []
    })

    for trip in categorized_trips:
        week_key = get_week_key(trip['started'])
        weekly_stats[week_key][trip['auto_category']] += trip['distance']
        weekly_stats[week_key]['total'] += trip['distance']
        weekly_stats[week_key]['trips'].append(trip)

        # Track Portland trips
        if is_portland_area(trip['start_address']) or is_portland_area(trip['end_address']):
            weekly_stats[week_key]['portland_miles'] += trip['distance']

        # Track Spokane trips
        if is_spokane_area(trip['start_address']) or is_spokane_area(trip['end_address']):
            weekly_stats[week_key]['spokane_miles'] += trip['distance']

        # Track weekend miles (Friday 5pm to Monday 6am)
        day_of_week = trip['started'].weekday()  # 0=Monday, 4=Friday, 6=Sunday
        hour = trip['started'].hour
        is_weekend = (day_of_week == 4 and hour >= 17) or \
                     (day_of_week == 5) or \
                     (day_of_week == 6) or \
                     (day_of_week == 0 and hour < 6)
        if is_weekend:
            weekly_stats[week_key]['weekend_miles'] += trip['distance']

    # Print weekly summary
    print("WEEKLY MILEAGE SUMMARY")
    print("-" * 80)
    print(f"{'Week Starting':<15} {'Commute':<12} {'Business':<12} {'Personal':<12} {'Total':<12}")
    print("-" * 80)

    total_commute = 0.0
    total_business = 0.0
    total_personal = 0.0
    total_all = 0.0

    for week in sorted(weekly_stats.keys()):
        stats = weekly_stats[week]
        commute = stats['commute']
        business = stats['business']
        personal = stats['personal']
        total = stats['total']

        print(f"{week:<15} {commute:>10.1f}mi {business:>10.1f}mi {personal:>10.1f}mi {total:>10.1f}mi")

        total_commute += commute
        total_business += business
        total_personal += personal
        total_all += total

    print("-" * 80)
    print(f"{'TOTAL':<15} {total_commute:>10.1f}mi {total_business:>10.1f}mi {total_personal:>10.1f}mi {total_all:>10.1f}mi")
    print()

    # Calculate business vs personal totals
    # Note: Commute is personal mileage
    actual_business = total_business
    actual_personal = total_personal + total_commute

    print("BUSINESS VS PERSONAL SUMMARY")
    print("-" * 80)
    print(f"Total Miles Driven:        {total_all:>10.1f} miles")
    print(f"Business Miles:            {actual_business:>10.1f} miles ({actual_business/total_all*100:.1f}%)")
    print(f"Personal Miles:            {actual_personal:>10.1f} miles ({actual_personal/total_all*100:.1f}%)")
    print(f"  - Commute:               {total_commute:>10.1f} miles")
    print(f"  - Other Personal:        {total_personal:>10.1f} miles")
    print()

    # Calculate special mileage totals
    total_portland = sum(stats.get('portland_miles', 0.0) for stats in weekly_stats.values())
    total_spokane = sum(stats.get('spokane_miles', 0.0) for stats in weekly_stats.values())
    total_weekend = sum(stats.get('weekend_miles', 0.0) for stats in weekly_stats.values())

    # Special mileage tracking
    print("SPECIAL MILEAGE TRACKING")
    print("-" * 80)
    print(f"Portland Round Trip Miles: {total_portland:>10.1f} miles")
    print(f"Spokane Round Trip Miles:  {total_spokane:>10.1f} miles")
    print(f"Weekend Miles (Fri 5pm-Mon 6am): {total_weekend:>10.1f} miles")
    print()

    # Detailed trip listing with business names
    print("=" * 80)
    print("DETAILED TRIP LOG")
    print("=" * 80)
    print()

    current_week = None
    for trip in categorized_trips:
        week = get_week_key(trip['started'])

        if week != current_week:
            if current_week is not None:
                print()
            print(f"Week of {week}")
            print("-" * 80)
            current_week = week

        date_str = trip['started'].strftime('%Y-%m-%d %H:%M')
        category = trip['auto_category'].upper()
        distance = trip['distance']
        business_name = trip['business_name']

        # Format: Date | Category | Distance | From -> To | Business Name
        from_addr = trip['start_address'][:30] if trip['start_address'] else 'Unknown'
        to_addr = trip['end_address'][:30] if trip['end_address'] else 'Unknown'

        print(f"{date_str} | {category:<8} | {distance:>6.1f}mi | {from_addr} -> {to_addr}")
        if category == 'BUSINESS':
            print(f"               Business: {business_name}")

    print()
    print("=" * 80)
    print("END OF REPORT")
    print("=" * 80)
    print()

    # Export to files
    if OPENPYXL_AVAILABLE:
        print("Exporting data to Excel file...")
        sys.stdout.flush()
        if export_to_excel(categorized_trips, weekly_stats, total_commute, total_business, total_personal, total_all):
            print("  [OK] ⭐ Excel report exported to: mileage_analysis.xlsx")
            print()
            print("💡 TIP: Open mileage_analysis.xlsx - all reports in one file with multiple sheets!")
            sys.stdout.flush()
        else:
            print("  [ERROR] Excel export failed, falling back to CSV...")
            sys.stdout.flush()
            export_csv_files(categorized_trips, weekly_stats, total_commute, total_business, total_personal, total_all)
    else:
        print("Exporting data to CSV files...")
        print("💡 NOTE: Install openpyxl for Excel export (pip install openpyxl)")
        sys.stdout.flush()
        export_csv_files(categorized_trips, weekly_stats, total_commute, total_business, total_personal, total_all)

    # Automatically analyze unresolved addresses after lookup
    if enable_lookup:
        print()
        print("-" * 80)
        print("Opening interactive address editor...")
        print("(Press Ctrl+C at any time to skip)")
        print("-" * 80)
        try:
            analyze_unresolved_addresses()
        except (EOFError, KeyboardInterrupt):
            print()
            print("Skipped.")
            print()
            pass

def export_csv_files(categorized_trips, weekly_stats, total_commute, total_business, total_personal, total_all):
    """Export all CSV files"""
    if export_comprehensive_report_csv(categorized_trips, weekly_stats):
        print("  [OK] Comprehensive report exported to: comprehensive_report.csv")
    if export_reconstructed_trips_csv(categorized_trips):
        print("  [OK] Reconstructed trips exported to: reconstructed_trips.csv")
    if export_weekly_summary_csv(weekly_stats):
        print("  [OK] Weekly summary exported to: weekly_summary.csv")
    if export_detailed_trips_csv(categorized_trips):
        print("  [OK] Detailed trips exported to: detailed_trips.csv")
    if export_summary_csv(total_commute, total_business, total_personal, total_all):
        print("  [OK] Summary exported to: summary.csv")
    if export_route_summary_csv(categorized_trips):
        print("  [OK] Route summary exported to: route_summary.csv")
    if export_address_business_correlation_csv(categorized_trips):
        print("  [OK] Address-business correlation exported to: address_business_correlation.csv")
    print()
    print("CSV export complete!")

def export_to_excel(categorized_trips, weekly_stats, total_commute, total_business, total_personal, total_all, filename="mileage_analysis.xlsx"):
    """Export all data to a single Excel file with multiple formatted sheets"""
    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Calculate totals
        total_portland = sum(stats.get('portland_miles', 0.0) for stats in weekly_stats.values())
        total_spokane = sum(stats.get('spokane_miles', 0.0) for stats in weekly_stats.values())
        total_weekend = sum(stats.get('weekend_miles', 0.0) for stats in weekly_stats.values())

        # Helper function to format headers
        def format_header_row(ws, row=1):
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Helper function to auto-size columns
        def auto_size_columns(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # SHEET 1: Summary
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(['MILEAGE ANALYSIS SUMMARY'])
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.append([])
        ws_summary.append(['Metric', 'Miles', 'Percentage'])
        format_header_row(ws_summary, 3)

        # Add data rows with proper values
        ws_summary.append(['Total Miles Driven', round(total_all, 2), 1.0])
        ws_summary.append(['Business Miles', round(total_business, 2), total_business/total_all])
        ws_summary.append(['Personal Miles (incl. Commute)', round(total_personal + total_commute, 2),
                          (total_personal + total_commute)/total_all])
        ws_summary.append(['  - Commute', round(total_commute, 2), total_commute/total_all])
        ws_summary.append(['  - Other Personal', round(total_personal, 2), total_personal/total_all])

        # Format number columns
        for row in range(4, 9):
            ws_summary[f'B{row}'].number_format = '#,##0.00'  # Miles with comma separator and 2 decimals
            ws_summary[f'C{row}'].number_format = '0.00%'  # Percentage with 2 decimals

        ws_summary.append([])
        ws_summary.append(['Special Tracking', 'Miles'])
        format_header_row(ws_summary, 10)
        ws_summary.append(['Portland Round Trip Miles', round(total_portland, 2)])
        ws_summary.append(['Spokane Round Trip Miles', round(total_spokane, 2)])
        ws_summary.append(['Weekend Miles (Fri 5pm-Mon 6am)', round(total_weekend, 2)])

        # Format special tracking numbers
        for row in range(11, 14):
            ws_summary[f'B{row}'].number_format = '#,##0.00'

        auto_size_columns(ws_summary)

        # SHEET 2: Weekly Breakdown
        ws_weekly = wb.create_sheet("Weekly Breakdown")
        ws_weekly.append(['Week Starting', 'Commute', 'Business', 'Personal', 'Total',
                         'Portland RT', 'Spokane RT', 'Weekend'])
        format_header_row(ws_weekly)

        start_row = 2
        for week in sorted(weekly_stats.keys()):
            stats = weekly_stats[week]
            ws_weekly.append([week,
                            round(stats['commute'], 2),
                            round(stats['business'], 2),
                            round(stats['personal'], 2),
                            round(stats['total'], 2),
                            round(stats.get('portland_miles', 0.0), 2),
                            round(stats.get('spokane_miles', 0.0), 2),
                            round(stats.get('weekend_miles', 0.0), 2)])

        total_row = start_row + len(weekly_stats)
        ws_weekly.append(['TOTAL',
                         round(total_commute, 2),
                         round(total_business, 2),
                         round(total_personal, 2),
                         round(total_all, 2),
                         round(total_portland, 2),
                         round(total_spokane, 2),
                         round(total_weekend, 2)])

        # Format all number columns (B through H)
        for row in range(start_row, total_row + 1):
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws_weekly[f'{col}{row}'].number_format = '#,##0.00'

        auto_size_columns(ws_weekly)

        # SHEET 3: Reconstructed Trips
        ws_recon = wb.create_sheet("Reconstructed Trips")
        ws_recon.append(['Date', 'Start Time', 'End Time', 'Duration (min)', 'Category',
                        'Start Address', 'End Address', 'Stops', 'Business Stops',
                        'Total Distance (miles)', 'Segment Count'])
        format_header_row(ws_recon)

        full_trips = reconstruct_full_trips(categorized_trips)
        recon_start_row = 2
        for trip in full_trips:
            duration = round((trip['end_time'] - trip['start_time']).total_seconds() / 60, 0)
            stops_str = ' -> '.join([s['address'] for s in trip['stops']]) if trip['stops'] else 'Direct'
            business_stops_str = '; '.join([f"{b['business_name']} ({b['distance']:.1f}mi)"
                                           for b in trip['business_stops']]) if trip['business_stops'] else ''

            ws_recon.append([trip['start_time'], trip['start_time'], trip['end_time'], duration,
                           trip['primary_category'].title(), trip['start_address'],
                           trip['end_address'], stops_str, business_stops_str,
                           round(trip['total_distance'], 2), len(trip['segments'])])

        # Format date and time columns
        recon_end_row = recon_start_row + len(full_trips) - 1
        for row in range(recon_start_row, recon_end_row + 1):
            ws_recon[f'A{row}'].number_format = 'yyyy-mm-dd'  # Date
            ws_recon[f'B{row}'].number_format = 'hh:mm'  # Start time
            ws_recon[f'C{row}'].number_format = 'hh:mm'  # End time
            ws_recon[f'D{row}'].number_format = '#,##0'  # Duration (whole number with commas)
            ws_recon[f'J{row}'].number_format = '#,##0.00'  # Distance with commas

        auto_size_columns(ws_recon)

        # SHEET 4: Detailed Trips
        ws_detailed = wb.create_sheet("Detailed Trips")
        ws_detailed.append(['Date', 'Time', 'Day of Week', 'Category', 'Distance (miles)',
                          'Start Address', 'End Address', 'Business Name',
                          'Start Odometer', 'End Odometer', 'Week Starting', 'Notes'])
        format_header_row(ws_detailed)

        detailed_start_row = 2
        for trip in categorized_trips:
            day_of_week = trip['started'].strftime('%A')
            category = trip['auto_category'].title()
            business_name = trip['business_name'] if trip['auto_category'] in ['business', 'commute'] else ''
            week = get_week_key(trip['started'])
            notes = trip.get('notes', '')  # Get notes if available

            ws_detailed.append([trip['started'], trip['started'], day_of_week, category, round(trip['distance'], 2),
                              trip['start_address'], trip['end_address'], business_name,
                              trip['start_odometer'], trip['end_odometer'], week, notes])

        # Format date, time, and distance columns
        detailed_end_row = detailed_start_row + len(categorized_trips) - 1
        for row in range(detailed_start_row, detailed_end_row + 1):
            ws_detailed[f'A{row}'].number_format = 'yyyy-mm-dd'  # Date
            ws_detailed[f'B{row}'].number_format = 'hh:mm'  # Time
            ws_detailed[f'E{row}'].number_format = '#,##0.00'  # Distance with commas

        auto_size_columns(ws_detailed)

        # SHEET 5: Address-Business Correlation
        ws_addr = wb.create_sheet("Address-Business Lookup")
        ws_addr.append(['Address', 'Business Name', 'Total Miles', 'Trip Count'])
        format_header_row(ws_addr)

        address_data = defaultdict(lambda: {'business_name': '', 'total_miles': 0.0, 'trip_count': 0})
        for trip in categorized_trips:
            if trip['auto_category'] == 'business' and trip['business_name']:
                end_addr = trip['end_address']
                if end_addr:
                    if not address_data[end_addr]['business_name']:
                        address_data[end_addr]['business_name'] = trip['business_name']
                    address_data[end_addr]['total_miles'] += trip['distance']
                    address_data[end_addr]['trip_count'] += 1

        sorted_addresses = sorted(address_data.items(), key=lambda x: x[1]['total_miles'], reverse=True)
        addr_start_row = 2
        for address, data in sorted_addresses:
            ws_addr.append([address, data['business_name'], round(data['total_miles'], 2), data['trip_count']])

        # Format number columns
        addr_end_row = addr_start_row + len(sorted_addresses) - 1
        for row in range(addr_start_row, addr_end_row + 1):
            ws_addr[f'C{row}'].number_format = '#,##0.00'  # Total miles with commas

        auto_size_columns(ws_addr)

        # Save workbook
        wb.save(filename)
        return True

    except Exception as e:
        print(f"Error exporting to Excel: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return False

def analyze_unresolved_addresses():
    """Analyze addresses that couldn't be resolved to business names"""

    # Initialize variables used throughout the function
    likely_businesses = []
    trip_analysis = {}

    print("=" * 80)
    print("UNRESOLVED ADDRESS ANALYSIS WITH SMART SUGGESTIONS")
    print("=" * 80)
    print()
    sys.stdout.flush()

    if not os.path.exists(BUSINESS_MAPPING_FILE):
        print("No business_mapping.json found!")
        print("Run the analysis with --lookup first to generate mappings.")
        return

    # Load mapping
    load_business_mapping()

    # Find all NO_BUSINESS_FOUND entries
    unresolved = [addr for addr in business_mapping.keys() if get_mapping_name(addr) == "NO_BUSINESS_FOUND"]
    resolved = {addr: get_mapping_name(addr) for addr in business_mapping.keys() if get_mapping_name(addr) != "NO_BUSINESS_FOUND"}

    print(f"Total addresses in mapping: {len(business_mapping)}")
    print(f"  - Resolved (with business names): {len(resolved)}")
    print(f"  - Unresolved (NO_BUSINESS_FOUND): {len(unresolved)}")
    print()
    sys.stdout.flush()

    if not unresolved:
        print("All addresses have been resolved!")
        print()
        return

    # === SMART AUTO-CATEGORIZATION ===
    print("-" * 80)
    print("SMART AUTO-CATEGORIZATION")
    print("-" * 80)
    print()
    sys.stdout.flush()

    # Patterns for residential street types
    residential_patterns = [
        r'\b(Ln|Lane)\b', r'\b(Ct|Court)\b', r'\b(Pl|Place)\b',
        r'\b(Way)\b', r'\b(Dr|Drive)\b', r'\b(Cir|Circle)\b',
        r'\b(Ter|Terrace)\b', r'\b(Loop)\b', r'\b(Rd)\b',
        r'Inglewood', r'Ross Rd'  # Known residential roads
    ]

    # Patterns for highways/routes (transit points - skip these)
    highway_patterns = [
        r'^(I-\d+|WA-\d+|SR-\d+|US-\d+)', r'\bHighway\b', r'\bHwy\b',
        r'^[\d]+ (I-|WA-|SR-|US-)', r'^WA-\d+', r'^I-\d+',
        r'Bothell Everett Hwy'  # Specific highway
    ]

    auto_personal = []
    auto_office = []
    auto_skip = []
    needs_review = []

    for addr in unresolved:
        # Check if it's a highway/route (skip)
        if any(re.search(p, addr, re.IGNORECASE) for p in highway_patterns):
            auto_skip.append(addr)
            continue

        # Check if it's near Office (same street, close number)
        nearby = find_nearby_addresses(addr, {WORK_ADDRESS: "Office", HOME_ADDRESS: "Home"}, max_number_diff=200)
        if nearby:
            auto_office.append((addr, nearby[0]['business_name'], nearby[0]['distance']))
            continue

        # Check if it's likely residential
        if any(re.search(p, addr, re.IGNORECASE) for p in residential_patterns):
            auto_personal.append(addr)
            continue

        # Needs manual review
        needs_review.append(addr)

    # Report findings
    print(f"Auto-categorization results:")
    print(f"  - Residential (auto-personal): {len(auto_personal)}")
    print(f"  - Near Office/Home (auto-match): {len(auto_office)}")
    print(f"  - Highways/Routes (skip): {len(auto_skip)}")
    print(f"  - Needs manual review: {len(needs_review)}")
    print()

    # Ask to apply auto-categorization
    if auto_personal or auto_office or auto_skip:
        print("Apply auto-categorization? This will:")
        if auto_personal:
            print(f"  - Mark {len(auto_personal)} residential addresses as PERSONAL")
        if auto_office:
            print(f"  - Mark {len(auto_office)} addresses as Office/Home (nearby)")
        if auto_skip:
            print(f"  - Skip {len(auto_skip)} highway/route addresses")
        print()

        try:
            response = input("Apply? (Y/n): ").strip().lower()
            if response != 'n' and response != 'no':
                # Apply auto-personal
                for addr in auto_personal:
                    business_mapping[addr] = "[PERSONAL]"

                # Apply auto-office
                for addr, match_name, dist in auto_office:
                    business_mapping[addr] = match_name

                # Save changes
                save_count = len(auto_personal) + len(auto_office)
                if save_count > 0:
                    try:
                        existing_data = {}
                        if os.path.exists(BUSINESS_MAPPING_FILE):
                            with open(BUSINESS_MAPPING_FILE, 'r', encoding='utf-8') as f:
                                existing_data = json.load(f)
                        for key, value in business_mapping.items():
                            existing_data[key] = value
                        with open(BUSINESS_MAPPING_FILE, 'w', encoding='utf-8') as f:
                            json.dump(existing_data, f, indent=2)
                        print()
                        print(f"✓ Auto-categorized {save_count} addresses!")
                        print(f"  - {len(auto_personal)} marked as personal")
                        print(f"  - {len(auto_office)} matched to Office/Home")
                    except Exception as e:
                        print(f"Error saving: {e}")

                # Update unresolved list to only show remaining
                unresolved = needs_review
                print()
                print(f"Remaining addresses needing manual review: {len(unresolved)}")
        except (EOFError, KeyboardInterrupt):
            print()
            print("Skipped auto-categorization.")

    print()
    sys.stdout.flush()

    if not unresolved:
        print("All addresses have been categorized!")
        print()
        return

    # Try to load trip data for smarter analysis
    trip_analysis = {}
    csv_file = "volvo-trips-log.csv"

    if os.path.exists(csv_file):
        print("Loading trip data for intelligent analysis...")
        sys.stdout.flush()
        try:
            trips = read_trips(csv_file)

            # Analyze each unresolved address in the context of trips
            for addr in unresolved:
                # Find all trips to/from this address
                addr_trips = [t for t in trips if addr in [t['start_address'], t['end_address']]]

                if addr_trips:
                    # Analyze visit patterns
                    weekday_visits = sum(1 for t in addr_trips if t['started'].weekday() < 5)
                    weekend_visits = len(addr_trips) - weekday_visits

                    # Business hours visits (8am-6pm weekdays)
                    business_hours = sum(1 for t in addr_trips
                                       if t['started'].weekday() < 5
                                       and 8 <= t['started'].hour < 18)

                    # Get date range
                    dates = [t['started'] for t in addr_trips]
                    first_visit = min(dates)
                    last_visit = max(dates)

                    trip_analysis[addr] = {
                        'visit_count': len(addr_trips),
                        'weekday_visits': weekday_visits,
                        'weekend_visits': weekend_visits,
                        'business_hours': business_hours,
                        'first_visit': first_visit,
                        'last_visit': last_visit,
                        'trips': addr_trips
                    }

            print(f"  [OK] Analyzed {len(trips)} trips for visit patterns")
            print()
            sys.stdout.flush()
        except Exception as e:
            print(f"  Note: Could not analyze trip data ({str(e)})")
            print("  Continuing with basic analysis...")
            print()
            sys.stdout.flush()

    # Group by city/area
    print("-" * 80)
    print("UNRESOLVED ADDRESSES BY AREA")
    print("-" * 80)
    print()
    sys.stdout.flush()

    by_area = defaultdict(list)
    for addr in sorted(unresolved):
        norm = normalize_address(addr)
        if "kenmore" in norm:
            by_area["Kenmore"].append(addr)
        elif "bothell" in norm:
            by_area["Bothell"].append(addr)
        elif any(x in norm for x in ["coupeville", "oak harbor", "clinton", "whidbey", "langley", "freeland"]):
            by_area["Whidbey Island"].append(addr)
        elif "seattle" in norm:
            by_area["Seattle"].append(addr)
        elif "everett" in norm:
            by_area["Everett"].append(addr)
        elif "mukilteo" in norm:
            by_area["Mukilteo"].append(addr)
        elif "portland" in norm:
            by_area["Portland"].append(addr)
        else:
            by_area["Other"].append(addr)

    for area in sorted(by_area.keys()):
        addrs = by_area[area]
        print(f"{area} ({len(addrs)} addresses):")
        for addr in sorted(addrs)[:10]:  # Show first 10
            print(f"  - {addr}")
        if len(addrs) > 10:
            print(f"  ... and {len(addrs) - 10} more")
        print()
        sys.stdout.flush()

    # Look for patterns - residential streets
    print("-" * 80)
    print("LIKELY RESIDENTIAL ADDRESSES")
    print("-" * 80)
    print()
    sys.stdout.flush()

    residential_patterns = ["ln ", "ct ", "pl ", "cir ", "way ", "dr ", "ave ", "st "]
    likely_residential = []

    for addr in sorted(unresolved):
        norm = normalize_address(addr)
        # Has a street number at the start and a residential street type
        if any(c.isdigit() for c in addr[:6]) and any(pattern in norm for pattern in residential_patterns):
            # Check if it looks like a residential address (not a business)
            if not any(keyword in norm for keyword in ["blvd", "hwy", "highway", "route", "state route"]):
                likely_residential.append(addr)

    print(f"Found {len(likely_residential)} likely residential addresses:")
    for addr in sorted(likely_residential)[:15]:
        print(f"  - {addr}")
    if len(likely_residential) > 15:
        print(f"  ... and {len(likely_residential) - 15} more")
    print()
    print("These are probably homes and may not need business names.")
    print()
    sys.stdout.flush()

    # Look for highways/routes
    print("-" * 80)
    print("HIGHWAYS & ROUTES")
    print("-" * 80)
    print()
    sys.stdout.flush()

    highways = []
    for addr in sorted(unresolved):
        norm = normalize_address(addr)
        if any(keyword in norm for keyword in ["i-90", "wa-", "hwy", "highway", "route", "state route", "fry"]):
            highways.append(addr)

    if highways:
        print(f"Found {len(highways)} highway/route addresses:")
        for addr in sorted(highways):
            print(f"  - {addr}")
        print()
        print("These may be rest stops or simply points on routes.")
        print()
        sys.stdout.flush()

    # Smart suggestions based on trip patterns
    if trip_analysis:
        print("-" * 80)
        print("SMART SUGGESTIONS - LIKELY BUSINESS ADDRESSES")
        print("-" * 80)
        print()
        print("Based on visit patterns, these addresses are likely businesses:")
        print()
        sys.stdout.flush()

        # Find addresses visited multiple times during business hours
        likely_businesses = []
        for addr, analysis in trip_analysis.items():
            # Heuristics for likely business:
            # - Visited 3+ times, OR
            # - Visited during business hours 2+ times, OR
            # - Visited on multiple weekdays
            score = 0
            reasons = []

            if analysis['visit_count'] >= 3:
                score += 3
                reasons.append(f"{analysis['visit_count']} visits")

            if analysis['business_hours'] >= 2:
                score += 2
                reasons.append(f"{analysis['business_hours']} weekday business hour visits")

            if analysis['weekday_visits'] >= 2:
                score += 1
                reasons.append(f"{analysis['weekday_visits']} weekday visits")

            # Check for suite/unit numbers (office building indicator)
            norm = normalize_address(addr)
            if any(x in norm for x in ['suite', 'ste', 'unit', '#']):
                score += 1
                reasons.append("has suite/unit number")

            if score >= 2:
                likely_businesses.append({
                    'address': addr,
                    'score': score,
                    'reasons': reasons,
                    'analysis': analysis
                })

        # Sort by score (highest first)
        likely_businesses.sort(key=lambda x: x['score'], reverse=True)

        if likely_businesses:
            for i, item in enumerate(likely_businesses[:20], 1):  # Show top 20
                addr = item['address']
                analysis = item['analysis']
                print(f"{i}. {addr}")
                print(f"   Confidence: {'★' * min(item['score'], 5)}")
                print(f"   Reasons: {', '.join(item['reasons'])}")
                print(f"   First visit: {analysis['first_visit'].strftime('%Y-%m-%d')}")
                print(f"   Last visit: {analysis['last_visit'].strftime('%Y-%m-%d')}")

                # Suggest a generic business name
                if analysis['visit_count'] >= 5:
                    print(f"   SUGGESTION: This is likely a regular customer/client location")
                    print(f"   ADD TO business_mapping.json as: \"Customer Site {chr(64+i)}\"")
                elif 'suite' in normalize_address(addr) or 'ste' in normalize_address(addr):
                    print(f"   SUGGESTION: Likely an office or professional services")
                    print(f"   ADD TO business_mapping.json as: \"Client Office - {addr.split(',')[1].strip() if ',' in addr else 'Unknown'}\"")
                else:
                    print(f"   SUGGESTION: Likely a business location")
                    print(f"   ADD TO business_mapping.json with a descriptive name")

                print()
                sys.stdout.flush()

            if len(likely_businesses) > 20:
                print(f"... and {len(likely_businesses) - 20} more likely businesses")
                print()
        else:
            print("No addresses with strong business patterns found.")
            print("Most unresolved addresses appear to be residential or one-time visits.")
            print()

        sys.stdout.flush()

    # Find potential fuzzy matches against already-resolved addresses
    print("-" * 80)
    print("FUZZY MATCH DETECTION")
    print("-" * 80)
    print()
    print("Checking for unresolved addresses similar to resolved ones...")
    print()
    sys.stdout.flush()

    suggested_mappings = {}
    for unresolved_addr in unresolved:
        norm_unresolved = normalize_address(unresolved_addr)

        # Check against all resolved addresses
        for resolved_addr, business_name in resolved.items():
            norm_resolved = normalize_address(resolved_addr)

            # Check if they're similar (one contains the other)
            if norm_unresolved in norm_resolved or norm_resolved in norm_unresolved:
                # Additional check: they should share significant parts
                unresolved_parts = set(norm_unresolved.split())
                resolved_parts = set(norm_resolved.split())
                common_parts = unresolved_parts & resolved_parts

                # If they share at least 3 parts, it's a match
                if len(common_parts) >= 3:
                    suggested_mappings[unresolved_addr] = {
                        'business_name': business_name,
                        'similar_to': resolved_addr
                    }
                    break

    if suggested_mappings:
        print(f"Found {len(suggested_mappings)} addresses that can be fuzzy matched:")
        print()

        for unresolved_addr, info in list(suggested_mappings.items())[:20]:
            print(f"Unresolved: {unresolved_addr}")
            print(f"  Similar to: {info['similar_to']}")
            print(f"  Business: {info['business_name']}")
            print()

        if len(suggested_mappings) > 20:
            print(f"... and {len(suggested_mappings) - 20} more")
            print()

        sys.stdout.flush()

        # Offer to add them to business_mapping.json
        print("-" * 80)
        print("AUTO-UPDATE BUSINESS MAPPINGS")
        print("-" * 80)
        print()
        print(f"Would you like to add these {len(suggested_mappings)} fuzzy matches to business_mapping.json?")
        print("This will allow them to be resolved in future analysis runs.")
        print()

        try:
            response = input("Add fuzzy matches? (y/N): ").strip().lower()

            if response == 'y':
                # Load existing business mapping
                load_business_mapping()

                # Add suggested mappings
                for unresolved_addr, info in suggested_mappings.items():
                    business_mapping[unresolved_addr] = info['business_name']

                # Save to file
                try:
                    # Read the file to preserve comments
                    existing_data = {}
                    if os.path.exists(BUSINESS_MAPPING_FILE):
                        with open(BUSINESS_MAPPING_FILE, 'r', encoding='utf-8') as f:
                            existing_data = json.load(f)

                    # Merge with new mappings
                    for key, value in business_mapping.items():
                        existing_data[key] = value

                    # Write back
                    with open(BUSINESS_MAPPING_FILE, 'w', encoding='utf-8') as f:
                        json.dump(existing_data, f, indent=2)

                    print()
                    print(f"✓ Added {len(suggested_mappings)} fuzzy matches to business_mapping.json")
                    print()
                    print("IMPORTANT: Run the analysis again (Option 2) to see these business names in your reports!")
                    print()
                    sys.stdout.flush()

                except Exception as e:
                    print()
                    print(f"Error saving mappings: {e}")
                    print()
                    sys.stdout.flush()
            else:
                print()
                print("Fuzzy matches not added. You can manually add them to business_mapping.json.")
                print()
                sys.stdout.flush()
        except (EOFError, KeyboardInterrupt):
            print()
            print("Skipping auto-update.")
            print()
            sys.stdout.flush()
    else:
        print("No potential fuzzy matches found.")
        print()
        sys.stdout.flush()

    # Interactive Address Editor
    print()
    print("-" * 80)
    print("INTERACTIVE ADDRESS EDITOR")
    print("-" * 80)
    print()
    print("QUICK GUIDE:")
    print("  y         = Use suggested nearby business")
    print("  m         = Open Google Maps")
    print("  p         = Mark as personal")
    print("  [name]    = Type business name")
    print("  [enter]   = Skip address")
    print("  Ctrl+C    = Exit editor")
    print()
    sys.stdout.flush()

    try:
        # Load existing business mapping
        load_business_mapping()

        # Track changes
        added_count = 0
        personal_count = 0

        # Go through unresolved addresses, prioritizing likely businesses
        addresses_to_edit = []

        # First add likely businesses (sorted by score)
        if likely_businesses:
            for item in likely_businesses:
                addresses_to_edit.append(('likely_business', item['address'], item))

        # Then add remaining unresolved addresses (sorted by visit count)
        remaining = []
        for addr in unresolved:
            # Skip if already in likely businesses
            if not any(lb['address'] == addr for lb in likely_businesses):
                visit_count = trip_analysis.get(addr, {}).get('visit_count', 0)
                remaining.append((addr, visit_count))
        remaining.sort(key=lambda x: x[1], reverse=True)

        for addr, _ in remaining:
            addresses_to_edit.append(('regular', addr, None))

        # Build list of all resolved addresses for proximity detection
        all_resolved = {}
        # Add home and office addresses first
        all_resolved[HOME_ADDRESS] = "Home"
        all_resolved[WORK_ADDRESS] = "Office"
        # Add from business mapping (unified storage)
        for addr in business_mapping.keys():
            name = get_mapping_name(addr)
            if name and name not in ["[PERSONAL]", "NO_BUSINESS_FOUND"]:
                all_resolved[addr] = name

        # Edit addresses
        for i, (addr_type, addr, extra_info) in enumerate(addresses_to_edit, 1):
            print(f"\n[{i}/{len(addresses_to_edit)}] {addr}")

            if addr_type == 'likely_business' and extra_info:
                analysis = extra_info['analysis']
                print(f"     → {analysis['visit_count']} visits, {analysis['business_hours']} during business hours")
                print(f"     → Likely BUSINESS (confidence: {extra_info['score']}/5)")
            elif addr in trip_analysis:
                analysis = trip_analysis[addr]
                print(f"     → {analysis['visit_count']} visits")

            # Check for nearby addresses
            nearby = find_nearby_addresses(addr, all_resolved)
            suggested_name = None
            if nearby:
                closest = nearby[0]
                print(f"     📍 NEARBY: {closest['business_name']} ({closest['distance']} numbers away)")
                print(f"         at {closest['address']}")
                suggested_name = closest['business_name']

            print()

            # Loop to allow "map" command
            while True:
                if suggested_name:
                    prompt = f"     [y={suggested_name} | m=map | p=personal | name | skip]: "
                else:
                    prompt = "     [m=map | p=personal | name | skip]: "

                name_input = input(prompt).strip()

                # Check if user wants to see map
                if name_input.lower() in ['map', 'm']:
                    print()
                    print("     🗺️  Opening Google Maps in your browser...")
                    try:
                        # Create Google Maps URL with the address
                        maps_url = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(addr)}"
                        webbrowser.open(maps_url)
                        print("     ✓ Map opened! Check your browser to see nearby businesses.")
                        print()
                    except Exception as e:
                        print(f"     ⚠️  Could not open browser: {e}")
                        print()
                    # Continue loop to re-prompt
                    continue

                # Check if user accepts the suggestion
                if suggested_name and name_input.lower() in ['y', 'yes']:
                    business_mapping[addr] = suggested_name
                    added_count += 1
                    print(f"     ✓ Using: {suggested_name}")
                    break

                # Process the actual input
                if name_input:
                    if name_input.lower() in ['p', 'personal']:
                        business_mapping[addr] = "[PERSONAL]"
                        personal_count += 1
                        print("     ✓ Marked as PERSONAL")
                    else:
                        business_mapping[addr] = name_input
                        added_count += 1
                        print(f"     ✓ Added: {name_input}")

                # Break out of while loop
                break

            sys.stdout.flush()

        # Save changes
        if added_count > 0 or personal_count > 0:
            print()
            print("-" * 80)
            print("SAVING CHANGES...")
            print("-" * 80)

            try:
                # Read existing file to preserve structure
                existing_data = {}
                if os.path.exists(BUSINESS_MAPPING_FILE):
                    with open(BUSINESS_MAPPING_FILE, 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)

                # Merge with new mappings
                for key, value in business_mapping.items():
                    existing_data[key] = value

                # Write back
                with open(BUSINESS_MAPPING_FILE, 'w', encoding='utf-8') as f:
                    json.dump(existing_data, f, indent=2)

                print()
                print(f"✓ Added {added_count} business names")
                print(f"✓ Marked {personal_count} addresses as personal")
                print()
                print("Changes saved to business_mapping.json")
                print()
                print("💡 TIP: Re-run analysis to see these names in your reports!")
                print()
                sys.stdout.flush()

            except Exception as e:
                print()
                print(f"ERROR saving changes: {e}")
                print()
                sys.stdout.flush()
        else:
            print()
            print("No changes made.")
            print()
            sys.stdout.flush()

    except (EOFError, KeyboardInterrupt):
        print()
        print("Editor cancelled.")
        print()
        sys.stdout.flush()

    # Recommendations
    print("-" * 80)
    print("RECOMMENDATIONS")
    print("-" * 80)
    print()
    print("1. LIKELY BUSINESSES (shown above):")
    print("   Add these to business_mapping.json with custom names.")
    print("   Example format:")
    print("   \"10484 Beardslee Blvd, Bothell\": \"ABC Company\"")
    print()
    print("2. RESIDENTIAL ADDRESSES:")
    print("   These don't need business names. The system handles them correctly.")
    print()
    print("3. ONE-TIME VISITS:")
    print("   Addresses visited only once may not need custom names unless")
    print("   you know they are important business locations.")
    print()
    print("4. USE FUZZY MATCHING:")
    print("   Partial addresses work! You can use just \"10484 Beardslee Blvd\"")
    print("   instead of the full \"10484 Beardslee Blvd, Bothell WA 98011\"")
    print()
    sys.stdout.flush()

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Mileage Analysis Tool for D\'Ewart Representatives, L.L.C.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  MileageAnalyzer.exe volvo-trips-log.csv
  MileageAnalyzer.exe volvo-trips-log.csv --lookup
  MileageAnalyzer.exe --analyze-unresolved

The script will generate:
  - Text report (to stdout)
  - mileage_analysis.xlsx (Excel file with all reports in multiple sheets)
        '''
    )

    parser.add_argument('csv_file', nargs='?', help='Path to the trip log CSV file')
    parser.add_argument('--lookup', '-l', action='store_true',
                       help='Enable online business lookup (requires geopy and internet connection)')
    parser.add_argument('--analyze-unresolved', '-a', action='store_true',
                       help='Analyze addresses that could not be resolved to business names')
    parser.add_argument('--start', '-s', help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end', '-e', help='End date (YYYY-MM-DD)')

    args = parser.parse_args()

    if args.analyze_unresolved:
        analyze_unresolved_addresses()
    elif args.csv_file:
        analyze_mileage(args.csv_file, enable_lookup=args.lookup,
                       start_date=args.start, end_date=args.end)
    else:
        parser.print_help()
